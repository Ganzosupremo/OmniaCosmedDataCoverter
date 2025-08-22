"""
Excel Exporter for COSMED data
Simplified version for web deployment
"""
import pandas as pd
from typing import List, Dict, Any
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelExporter:
    """
    Excel export functionality for COSMED data
    Simplified version for web deployment
    """
    
    # Standard selected parameters for COSMED analysis (15 key parameters)
    SELECTED_PARAMETERS = [
        "t", "Speed", "Pace", "VO2", "VO2/kg", "VCO2", 
        "METS", "RQ", "VE", "Rf", "HR", "VO2/HR",
        "P Syst", "P Diast", "HRR"
    ]
    
    def __init__(self, file_path: str):
        """Initialize Excel Exporter"""
        self.file_path = file_path
    
    def export_selected_parameters(self, extracted_data: List[Dict[str, Any]]) -> None:
        """Export selected parameters to Excel"""
        try:
            df = self._create_selected_parameters_dataframe(extracted_data)
            self._save_with_formatting(df, "Selected Parameters")
        except Exception as e:
            raise Exception(f"Error exporting selected parameters: {str(e)}")
    
    def export_max_values_only(self, extracted_data: List[Dict[str, Any]]) -> None:
        """Export max values only to Excel"""
        try:
            df = self._create_max_values_dataframe(extracted_data)
            self._save_with_formatting(df, "Max Values")
        except Exception as e:
            raise Exception(f"Error exporting max values: {str(e)}")
    
    def export_extracted_xml_data(self, extracted_data: List[Dict[str, Any]]) -> None:
        """Export complete extracted XML data to Excel"""
        try:
            df = self._create_complete_dataframe(extracted_data)
            self._save_with_formatting(df, "Complete Dataset")
        except Exception as e:
            raise Exception(f"Error exporting complete data: {str(e)}")
    
    def export_custom_parameters(self, extracted_data: List[Dict[str, Any]], custom_parameters: Dict[str, List[str]]) -> bool:
        """Export custom selected parameters to Excel"""
        try:
            df = self._create_custom_parameters_dataframe(extracted_data, custom_parameters)
            self._save_with_formatting(df, "Custom Parameters")
            return True
        except Exception as e:
            print(f"Error exporting custom parameters: {str(e)}")
            return False
    
    def _create_selected_parameters_dataframe(self, data: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create DataFrame with selected parameters"""
        rows = []
        
        for file_data in data:
            row = {
                'Filename': file_data.get('filename', ''),
                'Subject ID': file_data.get('subject_id', '')
            }
            
            # Convert parameters list to dictionary for easier lookup
            params_dict = {}
            for param in file_data.get('parameters', []):
                if param.get('Name'):
                    params_dict[param['Name']] = param
            
            # Add selected parameters
            for param_name in self.SELECTED_PARAMETERS:
                if param_name in params_dict:
                    param_data = params_dict[param_name]
                    unit = param_data.get('UM', '')
                    base_name = f"{param_name} ({unit})" if unit and unit != "---" else param_name
                    
                    # Special handling for VO2/kg - include multiple phases
                    if param_name == 'VO2/kg':
                        phases = ['MFO', 'AT', 'RC', 'Max']
                        for phase in phases:
                            value = param_data.get(phase, '')
                            if value:  # Only add if value exists
                                row[f"{base_name}_{phase}"] = value
                    else:
                        # For other parameters, use Max phase or Value phase
                        value = param_data.get('Max') or param_data.get('Value', '')
                        if value:  # Only add if value exists
                            row[f"{base_name}_Max"] = value
            
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _create_max_values_dataframe(self, data: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create DataFrame with max values only"""
        rows = []
        
        for file_data in data:
            row = {
                'Filename': file_data.get('filename', ''),
                'Subject ID': file_data.get('subject_id', '')
            }
            
            # Add all parameters with Max values
            for param in file_data.get('parameters', []):
                param_name = param.get('Name')
                if param_name:
                    max_value = param.get('Max')
                    if max_value:
                        unit = param.get('UM', '')
                        column_name = f"{param_name} ({unit})_Max" if unit and unit != "---" else f"{param_name}_Max"
                        row[column_name] = max_value
            
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _create_complete_dataframe(self, data: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create DataFrame with all parameter data"""
        rows = []
        
        for file_data in data:
            row = {
                'Filename': file_data.get('filename', ''),
                'Subject ID': file_data.get('subject_id', '')
            }
            
            # Add all parameters with all phases
            for param in file_data.get('parameters', []):
                param_name = param.get('Name')
                if param_name:
                    unit = param.get('UM', '')
                    base_name = f"{param_name} ({unit})" if unit and unit != "---" else param_name
                    
                    # Add all available phases
                    phases = ['Value', 'Rest', 'Warmup', 'MFO', 'AT', 'RC', 'Max', 'Pred', 'PercPred', 'Normal', 'Class']
                    for phase in phases:
                        value = param.get(phase)
                        if value:
                            row[f"{base_name}_{phase}"] = value
            
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _create_custom_parameters_dataframe(self, data: List[Dict[str, Any]], custom_parameters: Dict[str, List[str]]) -> pd.DataFrame:
        """Create DataFrame with custom selected parameters"""
        rows = []
        
        for file_data in data:
            row = {
                'Filename': file_data.get('filename', ''),
                'Subject ID': file_data.get('subject_id', '')
            }
            
            # Convert parameters list to dictionary for easier lookup
            params_dict = {}
            for param in file_data.get('parameters', []):
                if param.get('Name'):
                    params_dict[param['Name']] = param
            
            # Add custom selected parameters
            for param_name, selected_phases in custom_parameters.items():
                if param_name in params_dict:
                    param_data = params_dict[param_name]
                    unit = param_data.get('UM', '')
                    base_name = f"{param_name} ({unit})" if unit and unit != "---" else param_name
                    
                    for phase in selected_phases:
                        value = param_data.get(phase, '')
                        if value:  # Only add if value exists
                            # Use simple name for single phase, detailed name for multiple phases
                            if len(selected_phases) == 1:
                                column_name = base_name
                            else:
                                column_name = f"{base_name} - {phase}"
                            row[column_name] = value
            
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _save_with_formatting(self, df: pd.DataFrame, sheet_name: str = "COSMED Data"):
        """Save DataFrame to Excel with formatting"""
        try:
            with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Apply formatting
                self._apply_excel_formatting(workbook, worksheet, df)
                
        except Exception as e:
            # Fallback to simple save if formatting fails
            df.to_excel(self.file_path, index=False)
    
    def _apply_excel_formatting(self, workbook, worksheet, df):
        """Apply professional formatting to Excel worksheet"""
        try:
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply header formatting
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            print(f"Warning: Could not apply Excel formatting: {e}")
            # Continue without formatting