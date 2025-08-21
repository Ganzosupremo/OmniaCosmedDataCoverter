"""
Excel Formatter Module
Handles Excel file formatting and styling
"""
import pandas as pd
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelFormatter:
    """Excel formatting utilities for COSMED data"""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    def create_formatted_dataframe(self, data: List[Dict[str, Any]], data_type: str = "complete") -> pd.DataFrame:
        """
        Create formatted DataFrame from extracted data
        
        Args:
            data: List of extracted data dictionaries
            data_type: Type of data formatting ("complete", "selected", "max")
            
        Returns:
            Formatted pandas DataFrame
        """
        if not data:
            return pd.DataFrame()
        
        if data_type == "max":
            return self._create_max_values_dataframe(data)
        elif data_type == "selected":
            return self._create_selected_parameters_dataframe(data)
        else:
            return self._create_complete_dataframe(data)
    
    def _create_complete_dataframe(self, data: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create DataFrame with all parameters and phases"""
        rows = []
        
        for file_data in data:
            # Start with basic file information
            row = {
                'Filename': file_data['filename'],
                'Subject ID': file_data['subject_id'],
                'File Path': file_data['file_path']
            }
            
            # Add all parameters as columns
            for param in file_data['parameters']:
                param_name = param['Name']
                unit = param['UM']
                
                # Create base column name
                if unit and unit != "---":
                    base_name = f"{param_name} ({unit})"
                else:
                    base_name = param_name
                
                # Add all measurement phases
                phases = ['Value', 'Rest', 'Warmup', 'MFO', 'AT', 'RC', 'Max', 'Pred', 'PercPred', 'Normal', 'Class']
                for phase in phases:
                    column_name = f"{base_name}_{phase}"
                    row[column_name] = param.get(phase)
            
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _create_selected_parameters_dataframe(self, data: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create DataFrame with selected parameters only"""
        # Define the 15 key selected parameters
        selected_parameters = [
            "t", "Speed", "Pace", "VO2", "VO2/kg", "VCO2", 
            "METS", "RQ", "VE", "Rf", "HR", "VO2/HR"
        ]
        
        rows = []
        
        for file_data in data:
            row = {
                'Filename': file_data['filename'],
                'Subject ID': file_data['subject_id']
            }
            
            # Add selected parameters only
            for param in file_data['parameters']:
                param_name = param['Name']
                unit = param['UM']
                
                # Only process if this is a selected parameter
                if param_name in selected_parameters:
                    # Create parameter columns for specific phases
                    if unit and unit != "---":
                        base_name = f"{param_name} ({unit})"
                    else:
                        base_name = param_name
                    
                    # Define phases based on parameter - VO2/kg gets all phases, others get Max only
                    if param_name == 'VO2/kg':
                        phases_to_include = ['MFO', 'AT', 'RC', 'Max']
                    else:
                        phases_to_include = ['Max']
                    
                    for phase in phases_to_include:
                        if phase in param and param[phase] is not None:
                            column_name = f"{base_name}_{phase}"
                            row[column_name] = param[phase]
            
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _create_max_values_dataframe(self, data: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create DataFrame with only Max values"""
        rows = []
        
        for file_data in data:
            row = {
                'Filename': file_data['filename'],
                'Subject ID': file_data['subject_id']
            }
            
            # Add max parameters (already processed in data_extractor)
            if 'max_parameters' in file_data:
                row.update(file_data['max_parameters'])
            else:
                # Fallback: extract Max values from regular parameters
                for param in file_data.get('parameters', []):
                    param_name = param['Name']
                    unit = param['UM']
                    max_value = param.get('Max')
                    
                    if max_value is not None and max_value != '':
                        if unit and unit != "---":
                            column_name = f"{param_name} ({unit}) Max"
                        else:
                            column_name = f"{param_name} Max"
                        row[column_name] = max_value
            
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _get_relevant_phases_for_parameter(self, param_name: str) -> List[str]:
        """
        Get relevant measurement phases for a specific parameter
        
        Args:
            param_name: Name of the parameter
            
        Returns:
            List of relevant phases
        """
        # Define phase relevance based on parameter type
        if 'VO2' in param_name or 'VCO2' in param_name or 'VE' in param_name:
            return ['MFO', 'AT', 'RC', 'Max']
        elif 'HR' in param_name:
            return ['AT', 'RC', 'Max']
        else:
            # Default phases for other parameters
            return ['MFO', 'AT', 'RC', 'Max']
    
    def apply_excel_formatting(self, workbook: Workbook, worksheet_name: str = None) -> None:
        """
        Apply professional formatting to Excel workbook
        
        Args:
            workbook: Openpyxl workbook object
            worksheet_name: Name of worksheet to format (None for active)
        """
        if worksheet_name:
            ws = workbook[worksheet_name]
        else:
            ws = workbook.active
        
        # Format headers
        if ws.max_row > 0:
            for cell in ws[1]:  # First row
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.border
        
        # Apply borders to all data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = self.border
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, worksheet) -> None:
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set width with some padding, but cap at reasonable maximum
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_sheet(self, workbook: Workbook, data: List[Dict[str, Any]], sheet_name: str = "Summary") -> None:
        """
        Create a summary sheet with processing statistics
        
        Args:
            workbook: Openpyxl workbook object
            data: Processed data for statistics
            sheet_name: Name of summary sheet
        """
        ws = workbook.create_sheet(title=sheet_name)
        
        # Calculate statistics
        total_files = len(data)
        subjects_with_id = sum(1 for item in data if item.get('subject_id'))
        total_parameters = sum(len(item.get('parameters', [])) for item in data)
        unique_parameters = set()
        
        for item in data:
            for param in item.get('parameters', []):
                if param.get('Name'):
                    unique_parameters.add(param['Name'])
        
        # Create summary data
        summary_data = [
            ['Processing Summary', ''],
            ['Total Files Processed', total_files],
            ['Files with Subject ID', subjects_with_id],
            ['Total Parameters Extracted', total_parameters],
            ['Unique Parameter Types', len(unique_parameters)],
            ['', ''],
            ['Parameter Types Found', ''],
        ]
        
        # Add parameter list
        for param_name in sorted(unique_parameters):
            summary_data.append([param_name, ''])
        
        # Write data to sheet
        for row_idx, (key, value) in enumerate(summary_data, 1):
            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=value)
        
        # Format summary sheet
        ws.cell(1, 1).font = Font(bold=True, size=14)
        ws.cell(7, 1).font = Font(bold=True)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def save_formatted_excel(self, data: List[Dict[str, Any]], file_path: str, data_type: str = "complete") -> None:
        """
        Save data to formatted Excel file
        
        Args:
            data: Extracted data
            file_path: Output file path
            data_type: Type of formatting to apply
        """
        df = self.create_formatted_dataframe(data, data_type)
        
        # Create workbook with formatted data
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='COSMED Data', index=False)
            workbook = writer.book
            
            # Apply formatting
            self.apply_excel_formatting(workbook, 'COSMED Data')
            
            # Add summary sheet
            self.create_summary_sheet(workbook, data)
    
    def get_column_mapping(self, data_type: str) -> Dict[str, str]:
        """
        Get column name mapping for different export types
        
        Args:
            data_type: Type of export ("complete", "selected", "max")
            
        Returns:
            Dictionary mapping internal names to display names
        """
        base_mapping = {
            'filename': 'Filename',
            'subject_id': 'Subject ID',
            'file_path': 'File Path'
        }
        
        if data_type == "selected":
            # Add mappings for selected parameters
            selected_mapping = {
                'VO2/kg_MFO': 'VO2/kg (ml/min/kg) - MFO',
                'VO2/kg_AT': 'VO2/kg (ml/min/kg) - AT',
                'VO2/kg_RC': 'VO2/kg (ml/min/kg) - RC',
                'VO2/kg_Max': 'VO2/kg (ml/min/kg) - Max',
                # Add more as needed
            }
            base_mapping.update(selected_mapping)
        
        return base_mapping

    def create_custom_parameters_dataframe(self, data: List[Dict[str, Any]], custom_parameters: Dict[str, List[str]]) -> pd.DataFrame:
        """
        Create DataFrame with custom selected parameters and phases
        
        Args:
            data: List of extracted data dictionaries
            custom_parameters: Dictionary mapping parameter names to phases to export
                              e.g., {'VO2/kg': ['MFO', 'AT', 'RC', 'Max'], 'HR': ['Max']}
            
        Returns:
            Formatted pandas DataFrame with custom parameters
        """
        if not data:
            return pd.DataFrame()
        
        rows = []
        
        for file_data in data:
            row = {
                'Filename': file_data.get('filename', ''),
                'Subject ID': file_data.get('subject_id', ''),
            }
            
            # Get parameters data (it's a list of parameter dictionaries)
            parameters_list = file_data.get('parameters', [])
            
            # Convert parameters list to dictionary for easier access
            parameters_dict = {}
            for param in parameters_list:
                param_name = param.get('Name')
                if param_name:
                    parameters_dict[param_name] = param
            
            # Add custom selected parameters and phases
            for param_name, phases in custom_parameters.items():
                if param_name in parameters_dict:
                    param_data = parameters_dict[param_name]
                    
                    for phase in phases:
                        if phase in param_data:
                            # Create column name
                            if len(phases) == 1:
                                column_name = param_name
                            else:
                                column_name = f"{param_name} - {phase}"
                            
                            # Get the value from the parameter data
                            value = param_data.get(phase, '')
                            row[column_name] = value
                        else:
                            # Add empty value if phase not found
                            if len(phases) == 1:
                                column_name = param_name
                            else:
                                column_name = f"{param_name} - {phase}"
                            row[column_name] = ''
                else:
                    # Add empty values if parameter not found
                    for phase in phases:
                        if len(phases) == 1:
                            column_name = param_name
                        else:
                            column_name = f"{param_name} - {phase}"
                        row[column_name] = ''
            
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def save_custom_parameters_excel(self, data: List[Dict[str, Any]], file_path: str, custom_parameters: Dict[str, List[str]]) -> None:
        """
        Save custom parameters data to formatted Excel file
        
        Args:
            data: Extracted data
            file_path: Output file path
            custom_parameters: Dictionary mapping parameter names to phases to export
        """
        df = self.create_custom_parameters_dataframe(data, custom_parameters)
        
        # Create workbook with formatted data
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='COSMED Data', index=False)
            workbook = writer.book
            
            # Apply formatting
            self.apply_excel_formatting(workbook, 'COSMED Data')
            
            # Add custom summary sheet
            self.create_custom_summary_sheet(workbook, data, custom_parameters)
    
    def create_custom_summary_sheet(self, workbook: Workbook, data: List[Dict[str, Any]], custom_parameters: Dict[str, List[str]]) -> None:
        """
        Create summary sheet for custom parameters export
        
        Args:
            workbook: Excel workbook object
            data: Extracted data
            custom_parameters: Custom parameters configuration
        """
        summary_sheet = workbook.create_sheet('Export Summary')
        
        # Summary information
        summary_data = [
            ['Export Type', 'Custom Parameters'],
            ['Total Subjects', len(data)],
            ['Parameters Exported', len(custom_parameters)],
            ['Total Phases', sum(len(phases) for phases in custom_parameters.values())],
            [''],
            ['Parameter Configuration:'],
        ]
        
        # Add parameter details
        for param_name, phases in custom_parameters.items():
            summary_data.append([f'  {param_name}', ', '.join(phases)])
        
        # Write summary data
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Format headers
                if row_idx == 1 or cell_value == 'Parameter Configuration:':
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
        
        # Auto-adjust column widths
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_sheet.column_dimensions[column_letter].width = adjusted_width
